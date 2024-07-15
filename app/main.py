import concurrent.futures
import os
import platform
import subprocess
from typing import NamedTuple

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import (
    QFileDialog,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
)

CACHE_DIR = "./app/cache"


class SearchThread(QtCore.QThread):
    """Search for a keyword across Excel files in a folder"""

    # Signals for the status of the search process
    search_finished = QtCore.pyqtSignal(list)

    def __init__(self, folder: str, keyword: str):
        super().__init__()
        self.folder = folder
        self.keyword = keyword

    def search_row(self, file_path: str, sheet_name: str, row: NamedTuple):
        """Perform the search on a row of a sheet"""
        row_results = []
        for col_idx, cell in enumerate(row[1:], start=1):
            if pd.notna(cell) and self.keyword.lower() in str(cell).lower():
                cell_address = f"{get_column_letter(col_idx)}{row.Index + 2}"
                row_results.append((file_path, sheet_name, cell_address, str(cell)))
        return row_results

    def search_sheet(self, file_path: str, sheet_name: str):
        """Perform the search on a sheet of a file"""
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        sheet_results = []
        for row in df.itertuples():
            sheet_results.extend(self.search_row(file_path, sheet_name, row))
        return sheet_results

    def search_file(self, file_path: str):
        """Perform the search on a file"""
        excel_file = pd.ExcelFile(file_path)
        file_results = []
        for sheet_name in excel_file.sheet_names:
            file_results.extend(self.search_sheet(file_path, sheet_name))
        return file_results

    def run(self):
        """Perform the search by reading Excel files using the Pandas library"""
        results = []
        with concurrent.futures.ThreadPoolExecutor() as executor:
            futures = []
            for filename in os.listdir(self.folder):
                if filename.endswith((".xlsx", ".xls")):
                    file_path = os.path.join(self.folder, filename)
                    try:
                        futures.append(executor.submit(self.search_file, file_path))
                    except Exception as e:
                        print(f"Error reading {file_path}: {e}")
            # Collect the results
            for future in concurrent.futures.as_completed(futures):
                result = future.result()
                if result is not None:
                    results.extend(result)

        self.search_finished.emit(results)


class MoveThread(QtCore.QThread):
    """Use openpyxl to move the cursor to the cell address"""

    # Signals for the status of the move process
    move_finished = QtCore.pyqtSignal(str)

    def __init__(self, file_path, sheet_name, cell_address):
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.cell_address = cell_address

    def run(self):
        """Use openpyxl to move the cursor to the cell address"""
        try:
            wb = openpyxl.load_workbook(self.file_path)

            sheet = wb[self.sheet_name]
            wb.active = sheet
            sheet.sheet_view.selection[0].activeCell = self.cell_address
            sheet.sheet_view.selection[0].sqref = self.cell_address

            # Calculate row and column from the active cell
            row, col = coordinate_to_tuple(self.cell_address)
            # Display the cursor 10 rows from the top
            rows_offset = 10
            top_row = max(1, row - rows_offset)
            # Display the cursor 3 columns from the left
            cols_offset = 3
            left_col = max(1, col - cols_offset)
            top_left_cell = f"{get_column_letter(left_col)}{top_row}"
            sheet.sheet_view.topLeftCell = top_left_cell

            wb.save(self.file_path)

            self.move_finished.emit(self.file_path)
        except Exception as e:
            self.move_finished.emit(f"NG|{e}")


class ExcelSearchApp(QtWidgets.QWidget):
    """App GUI"""

    def __init__(self):
        super().__init__()
        self.folder_path: str = None
        self.search_thread: SearchThread = None
        self.move_thread = MoveThread
        self.initUI()

    def initUI(self):
        """Set up the layout of the app"""
        self.setWindowTitle("Excel Keyword Search")

        self.layout = QtWidgets.QVBoxLayout()

        self.folder_label = QtWidgets.QLabel("Select Folder:")
        self.layout.addWidget(self.folder_label)

        self.folder_button = QtWidgets.QPushButton("Browse")
        self.folder_button.clicked.connect(self.browse_folder)
        self.layout.addWidget(self.folder_button)

        self.keyword_label = QtWidgets.QLabel("Enter Keyword:")
        self.layout.addWidget(self.keyword_label)

        self.keyword_input = QtWidgets.QLineEdit()
        self.layout.addWidget(self.keyword_input)

        self.search_button = QtWidgets.QPushButton("Search")
        self.search_button.clicked.connect(self.start_search)
        self.layout.addWidget(self.search_button)

        self.results_table = QTableWidget()
        self.results_table.setColumnCount(5)
        self.results_table.setHorizontalHeaderLabels(
            ["File", "Sheet", "Cell", "Match", "Action"]
        )
        self.layout.addWidget(self.results_table)

        self.loading_label = QtWidgets.QLabel("Searching...")
        self.loading_label.setVisible(False)
        self.layout.addWidget(self.loading_label)

        self.opening_label = QtWidgets.QLabel("Opening...")
        self.opening_label.setVisible(False)
        self.layout.addWidget(self.opening_label)

        self.setLayout(self.layout)

    def browse_folder(self):
        """Let user choose the folder"""
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.folder_label.setText(f"Selected Folder: {folder}")
            self.folder_path = folder

    def start_search(self):
        """Trigger the search thread"""
        keyword = self.keyword_input.text()
        folder = getattr(self, "folder_path", None)

        if not folder:
            QMessageBox.warning(self, "Error", "Please select a folder.")
            return

        if not keyword:
            QMessageBox.warning(self, "Error", "Please enter a keyword.")
            return

        self.results_table.setRowCount(0)  # Clear previous results
        self.loading_label.setVisible(True)

        self.search_thread = SearchThread(folder, keyword)
        self.search_thread.search_finished.connect(self.display_results)
        self.search_thread.finished.connect(
            lambda: self.loading_label.setVisible(False)
        )
        self.search_thread.start()

    def display_results(self, results):
        """Render the search result as a table"""
        self.results_table.setRowCount(len(results))
        for row_idx, (file_path, sheet, addr, match) in enumerate(results):
            self.results_table.setItem(
                row_idx, 0, QTableWidgetItem(os.path.basename(file_path))
            )
            self.results_table.setItem(row_idx, 1, QTableWidgetItem(sheet))
            self.results_table.setItem(row_idx, 2, QTableWidgetItem(addr))
            self.results_table.setItem(row_idx, 3, QTableWidgetItem(match))

            open_file_button = QPushButton("Open")
            if file_path.endswith(".xls"):
                open_file_button.clicked.connect(
                    lambda checked, path=file_path, sheet=sheet, cell=addr: self.open_xls(
                        path, sheet, cell
                    )
                )
            else:
                open_file_button.clicked.connect(
                    lambda checked, path=file_path, sheet=sheet, cell=addr: self.open_xlsx(
                        path, sheet, cell
                    )
                )
            self.results_table.setCellWidget(row_idx, 4, open_file_button)

        self.results_table.setVisible(False)
        self.results_table.resizeColumnToContents(0)
        self.results_table.resizeColumnToContents(3)
        self.results_table.setVisible(True)

    def handle_move(self, signal: str):
        """Handle the move signal"""
        if signal.startswith("NG|"):
            QMessageBox.warning(self, "Error", f"Could not open file: {signal[3:]}")
        else:
            self.open_file_platform(signal)

    def open_file_platform(self, file_path):
        """Platform specific file opening"""
        if platform.system() == "Windows":
            os.startfile(file_path)
        elif platform.system() == "Darwin":  # macOS
            with subprocess.Popen(["open", file_path]) as instance:
                instance.wait()
        else:  # Linux and other
            with subprocess.Popen(["xdg-open", file_path]) as instance:
                instance.wait()

    def open_xlsx(self, file_path, sheet_name, cell_address):
        """Open the an XLSX file and move the cursor to the cell address"""
        try:
            self.opening_label.setVisible(True)
            self.move_thread = MoveThread(file_path, sheet_name, cell_address)
            self.move_thread.move_finished.connect(self.handle_move)
            self.move_thread.finished.connect(
                lambda: self.opening_label.setVisible(False)
            )
            self.move_thread.start()
        except Exception as e:
            self.opening_label.setVisible(False)
            QMessageBox.warning(self, "Error", f"Could not open file: {e}")

    def open_xls(self, file_path, sheet, cell):
        """Open an XLS file"""
        if not os.path.isdir(CACHE_DIR):
            os.mkdir(CACHE_DIR)

        basename = os.path.basename(file_path)
        xlsx_file_path = os.path.join(CACHE_DIR, basename[:-4] + ".xlsx")
        if not os.path.isfile(xlsx_file_path):
            reply = QMessageBox.warning(
                self,
                "Warning",
                "This will make an .xlsx copy out of your .xls and open the .xlsx version.\nAre you OK with it?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply == QMessageBox.Yes:
                df = pd.read_excel(file_path)
                df.to_excel(xlsx_file_path, index=False)
                self.open_xlsx(xlsx_file_path, sheet, cell)
        else:
            reply = QMessageBox.warning(
                self,
                "Warning",
                "Open the .xlsx version?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply == QMessageBox.Yes:
                self.open_xlsx(xlsx_file_path, sheet, cell)


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    ex = ExcelSearchApp()
    ex.resize(1024, 768)
    ex.show()
    sys.exit(app.exec_())
